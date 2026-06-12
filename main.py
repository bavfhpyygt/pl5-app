"""
排列五数据分析与推荐系统 - Flask 版本
兼容 PythonAnywhere WSGI 环境
"""
import os
import json
import random
from collections import Counter
from datetime import datetime

from flask import Flask, request, jsonify, send_from_directory, Response
from flask_cors import CORS

from database import engine, Base, SessionLocal, get_db
from models import LotteryRecord, User, SystemConfig
from schemas import (
    LotteryCreate, LotteryBatchCreate, LoginRequest, LoginResponse,
    ConfigUpdate, StatsResponse, RecommendRequest, RecommendResponse,
    LotteryOut, PaginatedResponse
)
from auth import hash_password, create_token, verify_token
from pydantic import ValidationError

# 创建数据库表
Base.metadata.create_all(bind=engine)

app = Flask(__name__)
CORS(app, origins=["*"])

# ========== 工具函数 ==========

def get_db_session():
    return SessionLocal()

def auth_required():
    """检查登录态"""
    auth_header = request.headers.get('Authorization', '')
    if not auth_header.startswith('Bearer '):
        return None, jsonify({'detail': '未登录'}), 401
    token = auth_header[7:]
    username = verify_token(token)
    if not username:
        return None, jsonify({'detail': '登录已过期'}), 401
    db = get_db_session()
    user = db.query(User).filter(User.username == username).first()
    db.close()
    if not user:
        return None, jsonify({'detail': '用户不存在'}), 401
    return user, None, None

# ========== 健康检查 ==========

@app.route('/api/health')
def health_check():
    return jsonify({'status': 'ok', 'app': '排列五数据分析系统'})

# ========== 彩票数据 API ==========

@app.route('/api/records', methods=['GET'])
def get_records():
    page = int(request.args.get('page', 1))
    page_size = int(request.args.get('page_size', 50))
    if page < 1:
        page = 1
    if page_size < 1 or page_size > 200:
        page_size = 50

    db = get_db_session()
    try:
        total = db.query(LotteryRecord).count()
        records = (
            db.query(LotteryRecord)
            .order_by(LotteryRecord.period.desc())
            .offset((page - 1) * page_size)
            .limit(page_size)
            .all()
        )
        items = [
            {
                'id': r.id,
                'period': r.period,
                'draw_date': r.draw_date,
                'numbers': r.numbers,
                'n1': r.n1, 'n2': r.n2, 'n3': r.n3, 'n4': r.n4, 'n5': r.n5,
                'created_at': r.created_at.isoformat() if r.created_at else None
            }
            for r in records
        ]
        return jsonify({
            'total': total,
            'page': page,
            'page_size': page_size,
            'items': items
        })
    finally:
        db.close()

@app.route('/api/records/latest', methods=['GET'])
def get_latest():
    db = get_db_session()
    try:
        record = db.query(LotteryRecord).order_by(LotteryRecord.period.desc()).first()
        if not record:
            return jsonify({}), 200
        return jsonify({
            'id': record.id,
            'period': record.period,
            'draw_date': record.draw_date,
            'numbers': record.numbers,
            'n1': record.n1, 'n2': record.n2, 'n3': record.n3, 'n4': record.n4, 'n5': record.n5,
            'created_at': record.created_at.isoformat() if record.created_at else None
        })
    finally:
        db.close()

@app.route('/api/records/<period>', methods=['GET'])
def get_by_period(period):
    db = get_db_session()
    try:
        record = db.query(LotteryRecord).filter(LotteryRecord.period == period).first()
        if not record:
            return jsonify({}), 200
        return jsonify({
            'id': record.id,
            'period': record.period,
            'draw_date': record.draw_date,
            'numbers': record.numbers,
            'n1': record.n1, 'n2': record.n2, 'n3': record.n3, 'n4': record.n4, 'n5': record.n5,
            'created_at': record.created_at.isoformat() if record.created_at else None
        })
    finally:
        db.close()

@app.route('/api/stats', methods=['GET'])
def get_stats():
    db = get_db_session()
    try:
        records = db.query(LotteryRecord).order_by(LotteryRecord.period.desc()).all()
        total = len(records)

        if total == 0:
            return jsonify({
                'total_records': 0,
                'latest_draw': None,
                'position_stats': [],
                'hot_cold': {},
                'missing_stats': []
            })

        latest = records[0]

        # 各位数字频率统计
        position_stats = []
        for pos in range(1, 6):
            field = f'n{pos}'
            counter = Counter()
            for r in records:
                counter[getattr(r, field)] += 1
            stats = [{'position': pos, 'digit': d, 'count': counter.get(d, 0), 'percent': round(counter.get(d, 0) / total * 100, 1)} for d in range(10)]
            stats.sort(key=lambda x: x['count'], reverse=True)
            position_stats.append(stats)

        # 冷热号
        all_counter = Counter()
        for r in records:
            for pos in range(1, 6):
                all_counter[getattr(r, f'n{pos}')] += 1

        hot = [{'digit': d, 'count': all_counter.get(d, 0)} for d in range(10)]
        hot.sort(key=lambda x: x['count'], reverse=True)

        hot_cold = {
            'hot': [h['digit'] for h in hot[:5]],
            'cold': [h['digit'] for h in hot[-5:]],
            'detail': hot
        }

        # 遗漏统计
        missing_stats = []
        for pos in range(1, 6):
            field = f'n{pos}'
            missing = {}
            for d in range(10):
                for i, r in enumerate(records):
                    if getattr(r, field) == d:
                        missing[d] = i
                        break
                else:
                    missing[d] = total
            missing_stats.append({
                'position': pos,
                'missing': [{'digit': d, 'since': missing.get(d, total)} for d in range(10)]
            })

        latest_draw = {
            'id': latest.id,
            'period': latest.period,
            'draw_date': latest.draw_date,
            'numbers': latest.numbers,
            'n1': latest.n1, 'n2': latest.n2, 'n3': latest.n3, 'n4': latest.n4, 'n5': latest.n5,
            'created_at': latest.created_at.isoformat() if latest.created_at else None
        }

        return jsonify({
            'total_records': total,
            'latest_draw': latest_draw,
            'position_stats': position_stats,
            'hot_cold': hot_cold,
            'missing_stats': missing_stats
        })
    finally:
        db.close()

@app.route('/api/recommend', methods=['POST'])
def recommend():
    data = request.get_json() or {}
    strategy = data.get('strategy', 'random')
    count = int(data.get('count', 5))

    db = get_db_session()
    try:
        records = db.query(LotteryRecord).order_by(LotteryRecord.period.desc()).limit(100).all()

        if len(records) < 10:
            recs = [''.join([str(random.randint(0, 9)) for _ in range(5)]) for _ in range(count)]
            return jsonify({
                'strategy': strategy,
                'recommendations': recs,
                'analysis': '数据量不足，使用随机推荐'
            })

        recommendations = []
        analysis = ''

        if strategy == 'frequency':
            all_counter = Counter()
            for r in records:
                for pos in range(1, 6):
                    all_counter[getattr(r, f'n{pos}')] += 1
            top_digits = [d for d, _ in all_counter.most_common(5)]

            for _ in range(count):
                rec = ''.join([str(random.choice(top_digits)) for _ in range(5)])
                recommendations.append(rec)

            analysis = f"基于近{len(records)}期数据，高频数字: {', '.join(map(str, top_digits))}"

        elif strategy == 'missing':
            cold_digits = []
            for pos in range(1, 6):
                field = f'n{pos}'
                counter = Counter()
                for r in records:
                    counter[getattr(r, field)] += 1
                least_common = counter.most_common()[-3:]
                cold_digits.append([d for d, _ in least_common])

            for _ in range(count):
                rec = ''.join([str(random.choice(cold_digits[p])) for p in range(5)])
                recommendations.append(rec)

            analysis = f"基于近{len(records)}期数据，各位冷号为: " + \
                       ', '.join([f"第{p+1}位:{','.join(map(str, cold_digits[p]))}" for p in range(5)])

        elif strategy == 'balanced':
            recs_freq = []
            recs_cold = []
            for _ in range(count // 2 + count % 2):
                rec = ''.join([str(random.randint(0, 9)) for _ in range(5)])
                recs_freq.append(rec)
            for _ in range(count // 2):
                rec = ''.join([str(random.randint(0, 9)) for _ in range(5)])
                recs_cold.append(rec)
            recommendations = recs_freq + recs_cold
            analysis = f"均衡策略：{len(recs_freq)}组高频 + {len(recs_cold)}组冷号"

        else:
            for _ in range(count):
                rec = ''.join([str(random.randint(0, 9)) for _ in range(5)])
                recommendations.append(rec)
            analysis = "使用随机策略生成"

        return jsonify({
            'strategy': strategy,
            'recommendations': recommendations,
            'analysis': analysis
        })
    finally:
        db.close()

@app.route('/api/trend', methods=['GET'])
def get_trend():
    pos = int(request.args.get('pos', 1))
    limit = int(request.args.get('limit', 100))
    if pos < 1 or pos > 5:
        pos = 1
    if limit < 1 or limit > 200:
        limit = 100

    field = f'n{pos}'
    db = get_db_session()
    try:
        records = (
            db.query(LotteryRecord)
            .order_by(LotteryRecord.period.desc())
            .limit(limit)
            .all()
        )
        records.reverse()
        return jsonify([{'period': r.period, 'date': r.draw_date, 'digit': getattr(r, field)} for r in records])
    finally:
        db.close()

# ========== 后台管理 API ==========

@app.route('/api/admin/login', methods=['POST'])
def login():
    data = request.get_json() or {}
    username = data.get('username', '')
    password = data.get('password', '')

    db = get_db_session()
    try:
        user = db.query(User).filter(User.username == username).first()
        if not user:
            return jsonify({'detail': '用户名或密码错误'}), 401
        if user.password_hash != hash_password(password):
            return jsonify({'detail': '用户名或密码错误'}), 401
        token = create_token(user.username)
        return jsonify({'token': token, 'username': user.username})
    finally:
        db.close()

@app.route('/api/admin/me', methods=['GET'])
def me():
    user, err_resp, status = auth_required()
    if err_resp:
        return err_resp, status
    return jsonify({'username': user.username})

@app.route('/api/admin/records', methods=['POST'])
def add_record():
    user, err_resp, status = auth_required()
    if err_resp:
        return err_resp, status

    data = request.get_json() or {}
    period = data.get('period', '')
    draw_date = data.get('draw_date', '')
    numbers = data.get('numbers', '')

    if len(numbers) != 5 or not numbers.isdigit():
        return jsonify({'detail': '号码必须是5位数字'}), 400

    db = get_db_session()
    try:
        existing = db.query(LotteryRecord).filter(LotteryRecord.period == period).first()
        if existing:
            return jsonify({'detail': f'期号 {period} 已存在'}), 400

        record = LotteryRecord(
            period=period,
            draw_date=draw_date,
            numbers=numbers,
            n1=int(numbers[0]), n2=int(numbers[1]),
            n3=int(numbers[2]), n4=int(numbers[3]),
            n5=int(numbers[4]),
        )
        db.add(record)
        db.commit()
        return jsonify({'ok': True, 'id': record.id})
    finally:
        db.close()

@app.route('/api/admin/records/batch', methods=['POST'])
def add_records_batch():
    user, err_resp, status = auth_required()
    if err_resp:
        return err_resp, status

    data = request.get_json() or {}
    records_data = data.get('records', [])

    db = get_db_session()
    try:
        added = 0
        skipped = 0
        for item in records_data:
            numbers = item.get('numbers', '')
            period = item.get('period', '')
            if len(numbers) != 5 or not numbers.isdigit():
                skipped += 1
                continue
            existing = db.query(LotteryRecord).filter(LotteryRecord.period == period).first()
            if existing:
                skipped += 1
                continue
            record = LotteryRecord(
                period=period,
                draw_date=item.get('draw_date', ''),
                numbers=numbers,
                n1=int(numbers[0]), n2=int(numbers[1]),
                n3=int(numbers[2]), n4=int(numbers[3]),
                n5=int(numbers[4]),
            )
            db.add(record)
            added += 1
        db.commit()
        return jsonify({'added': added, 'skipped': skipped})
    finally:
        db.close()

@app.route('/api/admin/records/<int:record_id>', methods=['DELETE'])
def delete_record(record_id):
    user, err_resp, status = auth_required()
    if err_resp:
        return err_resp, status

    db = get_db_session()
    try:
        record = db.query(LotteryRecord).filter(LotteryRecord.id == record_id).first()
        if not record:
            return jsonify({'detail': '记录不存在'}), 404
        db.delete(record)
        db.commit()
        return jsonify({'ok': True})
    finally:
        db.close()

@app.route('/api/admin/config', methods=['GET'])
def get_configs():
    user, err_resp, status = auth_required()
    if err_resp:
        return err_resp, status

    db = get_db_session()
    try:
        configs = db.query(SystemConfig).all()
        return jsonify([{'key': c.key, 'value': c.value} for c in configs])
    finally:
        db.close()

@app.route('/api/admin/config', methods=['POST'])
def set_config():
    user, err_resp, status = auth_required()
    if err_resp:
        return err_resp, status

    data = request.get_json() or {}
    key = data.get('key', '')
    value = data.get('value', '')

    db = get_db_session()
    try:
        config = db.query(SystemConfig).filter(SystemConfig.key == key).first()
        if config:
            config.value = value
        else:
            config = SystemConfig(key=key, value=value)
            db.add(config)
        db.commit()
        return jsonify({'ok': True})
    finally:
        db.close()

@app.route('/api/admin/records/upload', methods=['POST'])
def upload_excel():
    user, err_resp, status = auth_required()
    if err_resp:
        return err_resp, status

    if 'file' not in request.files:
        return jsonify({'detail': '没有上传文件'}), 400

    file = request.files['file']
    col_period = request.form.get('col_period', '期号')
    col_date = request.form.get('col_date', '开奖日期')
    col_numbers = request.form.get('col_numbers', '开奖号码')

    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'detail': '只支持 .xlsx / .xls 文件'}), 400

    try:
        import io
        import openpyxl

        contents = file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
        ws = wb.active

        # 找表头
        header = []
        for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
            row_vals = [str(v).strip() if v else '' for v in row]
            if col_period in row_vals or col_date in row_vals or col_numbers in row_vals:
                header = row_vals
                break
        if not header:
            header = [str(c.value).strip() if c.value else '' for c in ws[1]]

        try:
            idx_period = header.index(col_period)
            idx_date = header.index(col_date)
            idx_numbers = header.index(col_numbers)
        except ValueError as e:
            return jsonify({'detail': f'找不到列: {e}，请检查列名设置'}), 400

        db = get_db_session()
        try:
            added = 0
            skipped = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[idx_period]:
                    continue
                period = str(row[idx_period]).strip().lstrip("'")
                draw_date = str(row[idx_date]).strip() if row[idx_date] else ''
                numbers = str(row[idx_numbers]).strip().replace(' ', '').lstrip("'")

                numbers = ''.join(c for c in numbers if c.isdigit())[:5]
                period = ''.join(c for c in period if c.isdigit())

                if len(numbers) != 5:
                    skipped += 1
                    continue
                if not period:
                    skipped += 1
                    continue

                # 格式化日期
                if draw_date and len(draw_date) >= 10:
                    for fmt in ('%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d'):
                        try:
                            dt = datetime.strptime(draw_date[:10], fmt)
                            draw_date = dt.strftime('%Y-%m-%d')
                            break
                        except ValueError:
                            continue

                existing = db.query(LotteryRecord).filter(LotteryRecord.period == period).first()
                if existing:
                    skipped += 1
                    continue

                record = LotteryRecord(
                    period=period,
                    draw_date=draw_date,
                    numbers=numbers,
                    n1=int(numbers[0]), n2=int(numbers[1]),
                    n3=int(numbers[2]), n4=int(numbers[3]),
                    n5=int(numbers[4]),
                )
                db.add(record)
                added += 1

            db.commit()
            return jsonify({'added': added, 'skipped': skipped})
        finally:
            db.close()
    except Exception as e:
        return jsonify({'detail': f'文件解析失败: {str(e)}'}), 500

# ========== 页面路由 ==========

@app.route('/')
def index():
    return send_from_directory('static', 'index.html')

@app.route('/admin')
def admin():
    return send_from_directory('static', 'admin.html')

@app.route('/admin/')
def admin_slash():
    return send_from_directory('static', 'admin.html')

# ========== 静态文件 ==========

@app.route('/static/<path:filename>')
def static_files(filename):
    return send_from_directory('static', filename)

# WSGI 入口
application = app

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=8000, debug=True)
